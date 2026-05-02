"""Gestor de gastos: web en Flask con exportación/importación Excel."""

from __future__ import annotations

import json
import re
import uuid
from datetime import datetime
from io import BytesIO
from pathlib import Path

from flask import Flask, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

APP_DIR = Path(__file__).resolve().parent
DATA_PATH = APP_DIR / "data" / "transacciones.json"

app = Flask(__name__)


def load_rows() -> list[dict]:
    if not DATA_PATH.is_file():
        return []
    try:
        raw = json.loads(DATA_PATH.read_text(encoding="utf-8"))
        if isinstance(raw, list):
            return [r for r in raw if isinstance(r, dict)]
    except (json.JSONDecodeError, OSError):
        pass
    return []


def save_rows(rows: list[dict]) -> None:
    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    DATA_PATH.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")


def compute_saldos(rows: list[dict]) -> list[float]:
    acc = 0.0
    out: list[float] = []
    for r in rows:
        imp = float(r.get("importe", 0) or 0)
        if r.get("tipo") == "entrada":
            acc += imp
        else:
            acc -= imp
        out.append(acc)
    return out


def parse_fecha_key(fecha_str: str) -> tuple[int, int, int] | str:
    """Clave de calendario para agrupar por día; si no parsea, se usa el string tal cual."""
    s = (fecha_str or "").strip()
    if not s or s == "—":
        return "__sin_fecha__"
    parts = re.split(r"[/\-.]", s)
    parts = [p.strip() for p in parts if p.strip()]
    if len(parts) >= 2:
        try:
            d, m = int(parts[0]), int(parts[1])
            y = int(parts[2]) if len(parts) >= 3 else datetime.now().year
            if 1 <= m <= 12 and 1 <= d <= 31:
                return (y, m, d)
        except ValueError:
            pass
    return s


def compute_saldos_por_dia(rows: list[dict]) -> list[float]:
    """Saldo acumulado solo de movimientos del mismo día natural, en orden de la lista."""
    keys = [parse_fecha_key(str(r.get("fecha", ""))) for r in rows]
    out: list[float] = []
    for i in range(len(rows)):
        acc = 0.0
        k = keys[i]
        for j in range(i + 1):
            if keys[j] != k:
                continue
            imp = float(rows[j].get("importe", 0) or 0)
            if rows[j].get("tipo") == "entrada":
                acc += imp
            else:
                acc -= imp
        out.append(acc)
    return out


def normalize_header(cell: object) -> str:
    if cell is None:
        return ""
    return re.sub(r"\s+", " ", str(cell).strip().lower())


def parse_tipo(val: object) -> str | None:
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("entrada", "ingreso", "in"):
        return "entrada"
    if s in ("salida", "egreso", "gasto", "out"):
        return "salida"
    return None


def row_from_excel_cells(
    fecha, concepto, importe, tipo_raw
) -> dict | None:
    concepto_s = str(concepto).strip() if concepto is not None else ""
    if not concepto_s:
        return None
    try:
        if importe is None:
            return None
        imp = float(importe)
    except (TypeError, ValueError):
        return None
    if imp < 0:
        return None
    tipo = parse_tipo(tipo_raw)
    if tipo is None:
        return None
    if fecha is not None and hasattr(fecha, "strftime"):
        try:
            fecha_s = fecha.strftime("%d/%m/%Y")
        except (AttributeError, ValueError):
            fecha_s = str(fecha).strip()
    else:
        fecha_s = str(fecha).strip() if fecha is not None else ""
    if not fecha_s:
        fecha_s = "—"
    return {
        "id": str(uuid.uuid4()),
        "fecha": fecha_s,
        "concepto": concepto_s,
        "importe": imp,
        "tipo": tipo,
    }


def import_rows_from_workbook(wb) -> list[dict]:
    ws = wb.active
    rows_out: list[dict] = []
    header_map: dict[str, int] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        if i == 0:
            for idx, cell in enumerate(row):
                key = normalize_header(cell)
                if "fecha" in key:
                    header_map["fecha"] = idx
                elif "concepto" in key:
                    header_map["concepto"] = idx
                elif "importe" in key or "monto" in key:
                    header_map["importe"] = idx
                elif "entrada" in key or "salida" in key or "tipo" in key:
                    header_map["tipo"] = idx
            if len(header_map) < 4:
                return []
            continue
        try:
            f = row[header_map["fecha"]] if "fecha" in header_map else None
            c = row[header_map["concepto"]] if "concepto" in header_map else None
            im = row[header_map["importe"]] if "importe" in header_map else None
            t = row[header_map["tipo"]] if "tipo" in header_map else None
        except IndexError:
            continue
        parsed = row_from_excel_cells(f, c, im, t)
        if parsed:
            rows_out.append(parsed)
    return rows_out


def build_excel(rows: list[dict], saldos: list[float]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    header_fill = PatternFill("solid", fgColor="FFE8E8E8")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="FFC8C8C8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    saldos_dia = compute_saldos_por_dia(rows)
    headers = [
        "Fecha",
        "Concepto",
        "Importe",
        "Entrada / salida",
        "Saldo del día",
        "Saldo restante",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for i, (r, saldo_dia, saldo) in enumerate(
        zip(rows, saldos_dia, saldos), start=2
    ):
        ws.cell(row=i, column=1, value=r["fecha"]).alignment = Alignment(
            horizontal="center"
        )
        ws.cell(row=i, column=2, value=r["concepto"])
        ws.cell(row=i, column=3, value=float(r["importe"]))
        ws.cell(row=i, column=3).number_format = "#,##0"
        ws.cell(row=i, column=4, value=r["tipo"]).alignment = Alignment(
            horizontal="center"
        )
        ws.cell(row=i, column=5, value=float(saldo_dia))
        ws.cell(row=i, column=5).number_format = "#,##0"
        ws.cell(row=i, column=6, value=float(saldo))
        ws.cell(row=i, column=6).number_format = "#,##0"
        for c in range(1, 7):
            ws.cell(row=i, column=c).border = border
        ws.cell(row=i, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=i, column=5).alignment = Alignment(horizontal="right")
        ws.cell(row=i, column=6).alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


@app.route("/")
def index():
    rows = load_rows()
    saldos = compute_saldos(rows)
    saldos_dia = compute_saldos_por_dia(rows)
    items = list(zip(rows, saldos_dia, saldos))
    return render_template("index.html", items=items)


@app.post("/agregar")
def agregar():
    fecha = request.form.get("fecha", "").strip()
    concepto = request.form.get("concepto", "").strip()
    tipo = request.form.get("tipo", "entrada")
    try:
        importe = float(request.form.get("importe", "0").replace(",", "."))
    except ValueError:
        return redirect(url_for("index"))
    if not fecha or not concepto or importe < 0:
        return redirect(url_for("index"))
    if tipo not in ("entrada", "salida"):
        tipo = "entrada"
    rows = load_rows()
    rows.append(
        {
            "id": str(uuid.uuid4()),
            "fecha": fecha,
            "concepto": concepto,
            "importe": importe,
            "tipo": tipo,
        }
    )
    save_rows(rows)
    return redirect(url_for("index"))


@app.post("/eliminar/<row_id>")
def eliminar(row_id: str):
    rows = [r for r in load_rows() if r.get("id") != row_id]
    save_rows(rows)
    return redirect(url_for("index"))


@app.post("/limpiar")
def limpiar():
    save_rows([])
    return redirect(url_for("index"))


@app.get("/exportar.xlsx")
def exportar():
    rows = load_rows()
    saldos = compute_saldos(rows)
    bio = build_excel(rows, saldos)
    name = f"gastos_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/importar")
def importar():
    file = request.files.get("archivo")
    if not file or file.filename == "":
        return redirect(url_for("index"))
    reemplazar = request.form.get("reemplazar") == "1"
    try:
        data = file.read()
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        imported = import_rows_from_workbook(wb)
        wb.close()
    except Exception:
        return redirect(url_for("index"))
    if reemplazar:
        save_rows(imported)
    else:
        rows = load_rows()
        rows.extend(imported)
        save_rows(rows)
    return redirect(url_for("index"))


if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)
